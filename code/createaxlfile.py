from openpyxl import Workbook

wb=Workbook()
sheet=wb.active
sheet.title="Hcl"
'''
sheet["A1"].value=10
sheet["B2"].value="test"
sheet.cell(row=3,column=4).value="Hcl Data"
'''
'''
j=0
for i in range(10,60,10):
    j=j+1
    sheet.cell(row=j,column=1).value=i
    #for j in range(1,6):
        #sheet.cell(row=)
        '''
'''
for row in sheet.iter_rows(min_row=1,max_row=5,max_col=2,min_col=2):
    for r in row:
        r.value=1
        '''
j=0
for col in sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=5):

    for r in col:
        j = j + 100
        r.value=j


wb.save("C://Files/demoopenxlwrite.xlsx")